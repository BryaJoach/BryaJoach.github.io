from pathlib import Path
import openpyxl
import PySimpleGUI as sg
import pandas as pd
import numpy as np
import time

my_dict = {
    'A':0,
    'B':0,
    'C':0,
    'D':0,
    'E':0,
    'F':0,
    'G':0,
    'H':0,
    'I':0,
    'J':0,
    'K':0,
}


# Add some color to the window
sg.theme('DarkAmber')

year = time.strftime("%y", time.localtime())

# load the existing workbook
wb = openpyxl.load_workbook('QCRTemplate.xlsx')

# select the worksheet to write data to
ws = wb['Database']
Ticket_Count = len(ws['A'])
for x in range(1, Ticket_Count):
    ticket_ID = ws.cell(row=x+1, column=1)
    line_check = ticket_ID.value[0]

    my_dict[line_check] += 1 

layout = [

    [sg.Text('Please fill out the following fields:')],
    [sg.Text('Line', size=(15,1)), sg.InputText(key='Line')],
    [sg.Text('Month', size=(15,1)), sg.InputText(key='Month')],
    [sg.Text('Day', size=(15,1)), sg.InputText(key='Day')],
    [sg.Text('Shift', size=(15,1)), sg.Radio ('Morning', "Shift", default=True, size=(10,1), key='Shift'), sg.Radio('Afternoon', "Shift"), sg.Radio('Midnight', "Shift")],
    [sg.Text('Originator', size=(15,1)), sg.InputText(key='Originator')],
    [sg.Text('Bearing Part Number', size=(15,1)), sg.InputText(key='Bearing Part Number')],
    [sg.Text('WO#', size=(15,1)), sg.InputText(key='WO#')],
    [sg.Text('Machine', size=(15,1)), sg.Listbox(values=('FIN GOODS','ICG-A','ABK','LMU','AGS','KOG-A','ADM','ICG-B','OGF-A','ADS','AAM','AST','KOG-B','AMM','Inspection Table','AVS-1','AWC','FAS','APM','ACIS','ATS','ARSS','IGF-C','IAG-B','IGF-A','GS','IGF-B','IAG-A','AVS-2-A','Packing Table','IAG','W3','AVS-2-B','OGF-C','OGF-B','W1-OR','T118','T159','Bore Grind','Honer','OD grind','Width Grind','OD','After gauge','Bore','After Gauge','OGF #1','INT #1','GAD #1','Packing','B-Race #1','INT #2','GAD #2','AC','IGF #1','OGF #2','ATD #2','IGF #2','ASP2','B-Race #2','SVS','ASP1','AVG','ATD #1','AGEM','BUAM','ODC #1',
), size=(30, 3), key='Machine')],
    [sg.Text('Problem', size=(15,1)), sg.Listbox(values=('Ball','VCI','Large Radius','Small Radius','Radius','Roundness','Bearing','Roughness','Raceway ','Shoemark','Seal Lip','Seal Cut','Seal','Seal Dent','BP Mark','Bore','Reject','Clearance','No Etching','Packing','Taper NG','Master NG','Shape NG','LMU NG','Roughness NG','Sensor NG','NG','Oil','Missing Letter','Oil Stain','OD Dirt','OD Large','OD NG','OD Taper','OD Shoe Mark','OD','Gori','Off Center','Quantity','Dirt','Rust','Stain','Burn','Label','Grease','Broken/Damage','Other (Enter in Remarks)','Etching','Missing	OD','Shape','Mark','Shoe','Width','Mixed','Off','Damage','O/R','I/R','Chatter','Master','Raceway','Other','Radius NG','Seal','Burn','MASTER NOT REJECT','Burn Mark','Rejects','Rusty','Ball','Off centre','Shoe Mark','Drops','Wrong','Broken/Damaged','Taper'
), size=(30, 3), key='Problem')],
    [sg.Text('Employee Remarks',size=(15,1)), sg.Multiline(default_text='', size=(35, 3), key='Employee Remarks')],

    [sg.Submit(), sg.Button('Clear'), sg.Exit()]
]

window = sg.Window('RBB Submission Form', layout)

def clear_input():
    window['Line'].update('')
    window['Month'].update('')
    window['Day'].update('')
    window['Shift'].update('')
    window['Originator'].update('')
    window['Bearing Part Number'].update('')
    window['WO#'].update('')
    window['Employee Remarks'].update('')
    return None


while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'Exit':
        break
    if event == 'Clear':
        clear_input()
    if event == 'Submit':
        try:
            letter = values['Line']
            ticket_values = values
            my_dict[letter] +=1
            QCR = letter + "-" + year + "-" + str(my_dict[letter])
            data = [QCR, values['Line'], values['Month'], values['Day'], values['Shift'], values['Originator'], values['Bearing Part Number'], values['WO#'], values['Machine'][0], values['Problem'][0], values['Employee Remarks']]
            ws.append(data)
            wb.save('QCRTemplate.xlsx')
            sg.popup('Data saved!' + " " + QCR)
        except PermissionError:
            sg.popup('File currently being used')
window.close()
